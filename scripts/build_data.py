from __future__ import annotations

import argparse
import json
import math
import os
import re
import shutil
from collections import defaultdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / "public"
DATA_DIR = PUBLIC / "data"
PRICE_BOOK = ROOT / "各省光伏电价数据.xlsx"
CURVE_BOOK = ROOT / "光伏曲线.xlsx"
MAPPING_BOOK = ROOT / "各省价量映射维护表.xlsx"
LOCAL_CACHE_DIR = ROOT / ".data-cache"
PUBLIC_YEARLY_DIR = DATA_DIR / "yearly"

PROVINCE_ALIASES = {
    "冀南": ["河北南网", "冀南"],
    "蒙东": ["内蒙古东部", "蒙东"],
}

PROVINCE_DISPLAY_ALIASES = {
    "山东": "山东省",
    "山西": "山西省",
    "广东": "广东省",
    "广西": "广西壮族自治区",
    "云南": "云南省",
    "江苏": "江苏省",
    "湖北": "湖北省",
    "福建": "福建省",
    "安徽": "安徽省",
    "甘肃": "甘肃省",
    "湖南": "湖南省",
    "海南": "海南省",
    "贵州": "贵州省",
    "辽宁": "辽宁省",
    "吉林": "吉林省",
    "重庆": "重庆市",
    "陕西": "陕西省",
    "江西": "江西省",
}

PROVINCE_PRICE_PREFERENCES = {
    "广东": {
        "日前": ["日前统一结算点电价"],
        "实时": ["实际统一结算点电价"],
    },
    "湖南": {
        "日前": ["省内统一出清价格-日前"],
        "实时": ["省内统一出清价格-日内"],
    },
    "甘肃": {
        "日前": ["分区价格-统一结算点-日前电价", "统一结算点-日前电价"],
        "实时": ["分区价格-统一结算点-日内电价", "统一结算点-日内电价"],
    },
    "辽宁": {
        "日前": ["统一出清价格（调控后）", "统一出清价格(调控后)"],
        "实时": ["统一出清价格（调控后）", "统一出清价格(调控后)"],
    },
    "吉林": {
        "日前": ["统一出清价格-日前"],
        "实时": ["统一出清价格-日内"],
    },
}

PROVINCE_VOLUME_PREFERENCES = {
    "辽宁": {
        "日前": ["日前数据-新能源负荷-光伏", "新能源负荷-光伏"],
        "实时": ["实时数据-新能源负荷-光伏", "新能源负荷-光伏"],
    },
}

PROVINCES_WITHOUT_TYPICAL_VOLUME = {"甘肃"}

SEPARATE_PRICE_FILE_RULES = {
    "甘肃": ("分区节点电价",),
    "湖北": ("市场出清信息",),
    "福建": ("市场出清信息",),
    "吉林": ("省内出清价格",),
}


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value).strip()


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value) if math.isfinite(float(value)) else None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_date(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = cell_text(value)
    if not text:
        return None
    match = re.search(r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})", text)
    if match:
        y, m, d = map(int, match.groups())
        return f"{y:04d}-{m:02d}-{d:02d}"
    return None


def normalize_time(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    text = cell_text(value)
    if not text:
        return None
    match = re.search(r"(\d{1,2}):(\d{2})", text)
    if match:
        h, m = map(int, match.groups())
        return f"{h:02d}:{m:02d}"
    return None


def normalize_datetime(value: Any) -> tuple[str, str] | tuple[None, None]:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d"), value.strftime("%H:%M")
    text = cell_text(value)
    match = re.search(r"(20\d{2})[-/.年](\d{1,2})[-/.月](\d{1,2})[T\s]+(\d{1,2}):(\d{2})", text)
    if match:
        y, mo, d, h, mi = map(int, match.groups())
        return f"{y:04d}-{mo:02d}-{d:02d}", f"{h:02d}:{mi:02d}"
    return None, None


def month_from_filename(name: str) -> str | None:
    match = re.search(r"(20\d{2})[-_年 ]?(0[1-9]|1[0-2])[-_月 ]?\d{2}", name)
    if match:
        return f"{match.group(1)}-{match.group(2)}"
    return None


def flatten_headers(rows: list[tuple[Any, ...]]) -> tuple[list[str], int]:
    first = [cell_text(v) for v in rows[0]] if rows else []
    second = [cell_text(v) for v in rows[1]] if len(rows) > 1 else []
    has_second_header = bool(second) and any(v in ("日期", "时刻") for v in second[:3])
    if has_second_header:
        headers = []
        for a, b in zip(first, second):
            if a == b or not a:
                headers.append(b)
            elif not b:
                headers.append(a)
            else:
                headers.append(f"{a}-{b}")
        return headers, 3
    return first, 2


def find_col(headers: list[str], predicate) -> int | None:
    for idx, header in enumerate(headers):
        if predicate(header):
            return idx
    return None


def is_time_header(header: str) -> bool:
    return header in {"时刻", "时刻点"} or header.endswith("-时刻") or header.endswith("-时刻点")


def is_historical_file(path: Path) -> bool:
    return month_from_filename(path.name) is None and re.search(r"202[45]", path.name) is not None


def infer_market_from_filename(name: str) -> str | None:
    if "日前" in name:
        return "日前"
    if "实时" in name or "日内" in name:
        return "实时"
    return None


def load_workbook_for_path(path: Path):
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def prepare_sheet(ws) -> None:
    if hasattr(ws, "reset_dimensions"):
        ws.reset_dimensions()


def market_column_score(header: str, market: str) -> int:
    h = header.upper()
    if market == "日前":
        if "DAY_AHEAD" in h or "日前" in header or "预测" in header or "预平衡" in header:
            return 0
        return 9
    if "REAL_TIME" in h or "实时" in header or "日内" in header or "实际" in header:
        return 0
    return 9


def select_volume_columns(headers: list[str], province: str | None = None) -> dict[str, tuple[int, str, str]]:
    result: dict[str, tuple[int, str, str]] = {}
    explicit = [
        (i, h) for i, h in enumerate(headers)
        if ("光伏" in h or "PHOTOVOLTAIC" in h.upper())
        and not any(x in h for x in ("电价", "价格", "比例"))
    ]
    for market in ("日前", "实时"):
        if explicit:
            configured = PROVINCE_VOLUME_PREFERENCES.get(province or "", {}).get(market, [])
            configured_matches = [item for token in configured for item in explicit if token in item[1]]
            if configured_matches:
                idx, header = configured_matches[0]
                result[market] = (idx, header, "按省级维护说明使用光伏字段")
                continue
            market_specific = [item for item in explicit if market_column_score(item[1], market) == 0]
            neutral = [
                item for item in explicit
                if market_column_score(item[1], "日前") != 0 and market_column_score(item[1], "实时") != 0
            ]
            if market_specific:
                result[market] = (
                    market_specific[0][0],
                    market_specific[0][1],
                    "光伏字段；空值时复用另一市场光伏字段，仍为空则使用典型曲线",
                )
                continue
            if neutral:
                result[market] = (neutral[0][0], neutral[0][1], "光伏字段")
                continue
            counterpart = "实时" if market == "日前" else "日前"
            counterpart_specific = [item for item in explicit if market_column_score(item[1], counterpart) == 0]
            if counterpart_specific:
                idx, header = counterpart_specific[0]
                result[market] = (idx, header, f"无{market}光伏字段，复用{counterpart}光伏字段")
                continue
        result[market] = (-1, "典型光伏曲线", "无可用光伏字段，使用典型曲线独立权重")
    return result


def select_price_columns(headers: list[str]) -> dict[str, tuple[int, str]]:
    result: dict[str, tuple[int, str]] = {}
    for market in ("日前", "实时"):
        def in_market(h: str) -> bool:
            if market == "日前":
                return "日前" in h and not any(x in h for x in ("预测", "调控后"))
            return any(x in h for x in ("实时", "实际", "日内"))

        candidates = [
            (i, h) for i, h in enumerate(headers)
            if in_market(h) and any(x in h for x in ("出清价格", "统一结算点电价", "节点均价", "电价", "加权均价"))
        ]
        if candidates:
            result[market] = candidates[0]
    return result


def price_candidates(headers: list[str], market: str, province: str | None = None) -> list[tuple[int, str]]:
    def in_market(h: str) -> bool:
        if market == "日前":
            return "日前" in h
        return any(x in h for x in ("实时", "实际", "日内"))

    rows = [
        (i, h) for i, h in enumerate(headers)
        if in_market(h) and any(x in h or x in h.upper() for x in ("出清价格", "出清均价", "统一结算点电价", "节点均价", "电价", "加权均价", "PRICE"))
    ]
    configured = PROVINCE_PRICE_PREFERENCES.get(province or "", {}).get(market, [])
    configured_rows = [x for token in configured for x in rows if token in x[1]]
    preferred = [x for x in rows if not any(bad in x[1] for bad in ("预测", "调控后"))]
    return configured_rows + [x for x in preferred if x not in configured_rows] + [x for x in rows if x not in configured_rows and x not in preferred]


def neutral_price_candidates(
    headers: list[str],
    market: str | None = None,
    province: str | None = None,
) -> list[tuple[int, str]]:
    rows = [
        (i, h) for i, h in enumerate(headers)
        if any(x in h or x in h.upper() for x in ("出清价格", "出清均价", "统一结算点电价", "节点均价", "节点电价", "电价", "加权均价", "PRICE"))
        and not any(bad in h for bad in ("比例", "类型", "日期", "时刻", "预测", "申报"))
    ]
    configured = PROVINCE_PRICE_PREFERENCES.get(province or "", {}).get(market or "", [])
    configured_rows = [x for token in configured for x in rows if token in x[1]]
    preferred = [
        x for x in rows
        if any(good in x[1] for good in ("出清", "节点", "统一")) and "调控后" not in x[1]
    ]
    return configured_rows + [x for x in preferred if x not in configured_rows] + [
        x for x in rows if x not in configured_rows and x not in preferred
    ]


def market_from_text(value: Any) -> str | None:
    text = cell_text(value)
    if "日前" in text:
        return "日前"
    if "实时" in text or "日内" in text:
        return "实时"
    return None


def row_price(
    row: tuple[Any, ...],
    headers: list[str],
    market: str,
    inferred_market: str | None = None,
    province: str | None = None,
) -> tuple[float | None, str | None]:
    configured = PROVINCE_PRICE_PREFERENCES.get(province or "", {}).get(market, [])
    if inferred_market == market and configured:
        configured_neutral = [
            item for item in neutral_price_candidates(headers, market, province)
            if any(token in item[1] for token in configured)
        ]
        for idx, header in configured_neutral:
            value = to_float(row[idx] if idx < len(row) else None)
            if value is not None:
                return value, header
    for idx, header in price_candidates(headers, market, province):
        value = to_float(row[idx] if idx < len(row) else None)
        if value is not None:
            return value, header
    if inferred_market == market:
        for idx, header in neutral_price_candidates(headers, market, province):
            value = to_float(row[idx] if idx < len(row) else None)
            if value is not None:
                return value, header
    return None, None


def row_volume(
    row: tuple[Any, ...],
    idx: int,
    t: str,
    typical_curve: dict[str, float],
    fallback_idx: int = -1,
    allow_typical: bool = True,
) -> tuple[float | None, bool]:
    if idx < 0:
        if not allow_typical:
            return None, False
        return typical_curve.get(t, 0), True
    value = to_float(row[idx] if idx < len(row) else None)
    if value is None and fallback_idx >= 0:
        value = to_float(row[fallback_idx] if fallback_idx < len(row) else None)
    if value is None:
        if not allow_typical:
            return None, False
        return typical_curve.get(t, 0), True
    return value, False


def selected_price_columns(
    headers: list[str],
    province: str | None = None,
    inferred_market: str | None = None,
) -> dict[str, str]:
    result = {}
    for market in ("日前", "实时"):
        candidates = price_candidates(headers, market, province)
        configured = PROVINCE_PRICE_PREFERENCES.get(province or "", {}).get(market, [])
        if inferred_market == market and configured:
            configured_neutral = [
                item for item in neutral_price_candidates(headers, market, province)
                if any(token in item[1] for token in configured)
            ]
            candidates = configured_neutral + [item for item in candidates if item not in configured_neutral]
        if not candidates:
            candidates = neutral_price_candidates(headers, market, province)
        if candidates:
            result[market] = " / ".join(header for _, header in candidates[:3])
    return result


def sheet_headers(ws) -> list[str]:
    prepare_sheet(ws)
    return [cell_text(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]


def ingest_split_market_workbook(path: Path, province: str, typical_curve: dict[str, float]) -> tuple[list[dict[str, Any]], dict[str, Any]] | None:
    wb = load_workbook_for_path(path)
    if "日前" not in wb.sheetnames or "实时" not in wb.sheetnames:
        return None
    day_ws = wb["日前"]
    real_ws = wb["实时"]
    prepare_sheet(day_ws)
    prepare_sheet(real_ws)
    day_headers = sheet_headers(day_ws)
    real_headers = sheet_headers(real_ws)
    if not day_headers or day_headers[0].lower() != "datetime":
        return None

    day_volume_col = find_col(day_headers, lambda h: "PHOTOVOLTAIC" in h.upper() or "光伏" in h)
    day_price_col = find_col(day_headers, lambda h: "PRICE" in h.upper() or "价格" in h or "电价" in h)
    real_price_col = find_col(real_headers, lambda h: "PRICE" in h.upper() or "价格" in h or "电价" in h)
    if day_volume_col is None or day_price_col is None:
        return None

    records: list[dict[str, Any]] = []
    used_typical = False
    day_volumes: dict[str, float] = {}
    for row in day_ws.iter_rows(min_row=2, values_only=True):
        d, t = normalize_datetime(row[0] if row else None)
        if not d or not t:
            continue
        dt = f"{d}T{t}:00"
        volume = to_float(row[day_volume_col] if day_volume_col < len(row) else None)
        if volume is None:
            volume = typical_curve.get(t, 0)
            used_typical = True
        day_volumes[dt] = volume
        price = to_float(row[day_price_col] if day_price_col < len(row) else None)
        if volume is not None and price is not None:
            records.append({
                "province": province,
                "month": d[:7],
                "datetime": dt,
                "market": "日前",
                "price": price,
                "volume": volume,
            })

    if real_price_col is not None:
        for row in real_ws.iter_rows(min_row=2, values_only=True):
            d, t = normalize_datetime(row[0] if row else None)
            if not d or not t:
                continue
            dt = f"{d}T{t}:00"
            price = to_float(row[real_price_col] if real_price_col < len(row) else None)
            volume = day_volumes.get(dt)
            if volume is None:
                volume = typical_curve.get(t, 0)
                used_typical = True
            if volume is not None and price is not None:
                records.append({
                    "province": province,
                    "month": d[:7],
                    "datetime": dt,
                    "market": "实时",
                    "price": price,
                    "volume": volume,
                })
    return records, {
        "file": path.name,
        "sheet": "日前/实时",
        "priceColumns": {
            "日前": day_headers[day_price_col],
            "实时": real_headers[real_price_col] if real_price_col is not None else "",
        },
        "volumeColumns": {
            "日前": day_headers[day_volume_col],
            "实时": day_headers[day_volume_col],
        },
        "volumeSource": {
            "日前": "光伏字段；为空点使用典型曲线独立权重",
            "实时": "实时表缺光伏字段，按日期时刻复用日前光伏字段；仍缺失时使用典型曲线",
        },
        "records": len(records),
        "usesTypicalCurve": used_typical,
    }


def load_typical_curves() -> dict[str, dict[str, float]]:
    wb = openpyxl.load_workbook(CURVE_BOOK, data_only=True)
    curves = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        curve = {}
        for col in range(1, ws.max_column + 1):
            slot = normalize_time(ws.cell(1, col).value)
            value = to_float(ws.cell(2, col).value)
            if slot is not None and value is not None:
                curve[slot] = value
        if curve:
            curves[sheet] = curve
    return curves


def ensure_curve_sheets(provinces_requiring_curve: set[str]) -> None:
    if not provinces_requiring_curve:
        return
    backup = CURVE_BOOK.with_suffix(".backup.xlsx")
    if not backup.exists():
        shutil.copy2(CURVE_BOOK, backup)
    wb = openpyxl.load_workbook(CURVE_BOOK)
    source = wb["通用"]
    changed = False
    for province in sorted(provinces_requiring_curve):
        if province in wb.sheetnames:
            continue
        ws = wb.copy_worksheet(source)
        ws.title = province
        changed = True
    if changed:
        wb.save(CURVE_BOOK)


def load_params() -> dict[str, dict[str, Any]]:
    wb = openpyxl.load_workbook(PRICE_BOOK, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header1 = [cell_text(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    header2 = [v for v in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    rows = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        province = cell_text(row[2] if len(row) > 2 else "")
        if not province:
            continue
        params = {
            "region": cell_text(row[1] if len(row) > 1 else ""),
            "coalBenchmark2025": to_float(row[3] if len(row) > 3 else None),
            "mechanism": {},
            "settlement": {},
        }
        for idx, label in enumerate(header1):
            year_match = re.search(r"(20\d{2})年机制竞价", label)
            if year_match and idx < len(row):
                year = year_match.group(1)
                params["mechanism"].setdefault(year, {})["price"] = to_float(row[idx])
            if "执行比例" in label and idx < len(row):
                year_match = re.search(r"(20\d{2})年", label)
                if year_match:
                    year = year_match.group(1)
                    params["mechanism"].setdefault(year, {})["ratio"] = to_float(row[idx])
        for idx, month_cell in enumerate(header2):
            if isinstance(month_cell, (datetime, date)) and idx < len(row):
                params["settlement"][month_cell.strftime("%Y-%m")] = to_float(row[idx])
        rows[province] = params
    return rows


def match_params(province: str, params: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    choices = PROVINCE_ALIASES.get(province, [])
    display = PROVINCE_DISPLAY_ALIASES.get(province, province)
    choices.extend([display, province])
    for choice in choices:
        if choice in params:
            return params[choice]
    for key, value in params.items():
        if province in key or key in province:
            return value
    return None


def ingest_standard_sheet(path: Path, province: str, typical_curve: dict[str, float]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook_for_path(path)
    ws = wb[wb.sheetnames[0]]
    prepare_sheet(ws)
    rows_iter = ws.iter_rows(values_only=True)
    header_rows = []
    for _ in range(2):
        try:
            header_rows.append(next(rows_iter))
        except StopIteration:
            break
    headers, start_row = flatten_headers(header_rows)
    data_rows = ws.iter_rows(min_row=start_row, values_only=True)

    date_col = find_col(headers, lambda h: h == "日期" or h.endswith("-日期"))
    time_col = find_col(headers, is_time_header)
    if date_col is None or time_col is None:
        return [], {"file": path.name, "ignored": "未识别日期/时刻列"}

    type_col = find_col(headers, lambda h: h == "类型" or h.endswith("-类型") or h == "电价类型")
    inferred_file_market = infer_market_from_filename(path.name)
    volume_cols = select_volume_columns(headers, province)
    records: list[dict[str, Any]] = []
    needs_curve = any(source.endswith("典型光伏曲线") for _, _, source in volume_cols.values())

    for row in data_rows:
        d = normalize_date(row[date_col] if date_col < len(row) else None)
        t = normalize_time(row[time_col] if time_col < len(row) else None)
        if not d or not t:
            continue
        month = d[:7]
        row_market = market_from_text(row[type_col] if type_col is not None and type_col < len(row) else None)
        markets = [row_market] if row_market else [inferred_file_market] if inferred_file_market else ["日前", "实时"]
        for market in markets:
            if market not in volume_cols:
                continue
            vidx, volume_header, volume_source = volume_cols[market]
            price, price_header = row_price(row, headers, market, row_market or inferred_file_market, province)
            other_market = "实时" if market == "日前" else "日前"
            fallback_idx = volume_cols.get(other_market, (-1, "", ""))[0]
            if fallback_idx == vidx:
                fallback_idx = -1
            volume, used_typical = row_volume(
                row,
                vidx,
                t,
                typical_curve,
                fallback_idx,
                province not in PROVINCES_WITHOUT_TYPICAL_VOLUME,
            )
            if used_typical:
                needs_curve = True
            if price is None or volume is None:
                continue
            records.append({
                "province": province,
                "month": month,
                "datetime": f"{d}T{t}:00",
                "market": market,
                "price": price,
                "volume": volume,
            })
    return records, {
        "file": path.name,
        "sheet": ws.title,
        "priceColumns": selected_price_columns(headers, province, inferred_file_market),
        "volumeColumns": {k: v[1] for k, v in volume_cols.items()},
        "volumeSource": {k: v[2] for k, v in volume_cols.items()},
        "records": len(records),
        "usesTypicalCurve": needs_curve,
    }


def ingest_southern_price_sheet(path: Path, province: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    wb = load_workbook_for_path(path)
    ws = wb[wb.sheetnames[0]]
    prepare_sheet(ws)
    rows = ws.iter_rows(values_only=True)
    headers = [cell_text(v) for v in next(rows)]
    records = []
    region_col = find_col(headers, lambda h: h == "区域")
    date_col = find_col(headers, lambda h: h == "日期")
    time_col = find_col(headers, is_time_header)
    type_col = find_col(headers, lambda h: h in {"电价类型", "类型"})
    neutral_prices = neutral_price_candidates(headers)
    for row in rows:
        region = cell_text(row[region_col]) if region_col is not None and region_col < len(row) else province
        if region and province not in region:
            continue
        if "datetime" in headers[0].lower():
            dt_text = cell_text(row[0])
            d = dt_text[:10]
            t = dt_text[11:16]
            offset = 1
        else:
            d = normalize_date(row[date_col]) if date_col is not None and date_col < len(row) else None
            t = normalize_time(row[time_col]) if time_col is not None and time_col < len(row) else None
            offset = 0
        if not d or not t:
            continue
        row_market = market_from_text(row[type_col]) if type_col is not None and type_col < len(row) else None
        if row_market:
            for idx, header in neutral_prices:
                price = to_float(row[idx] if idx < len(row) else None)
                if price is not None:
                    records.append({
                        "province": province,
                        "month": d[:7],
                        "datetime": f"{d}T{t}:00",
                        "market": row_market,
                        "price": price,
                        "volume": None,
                    })
                    break
            continue
        for idx, header in enumerate(headers):
            price = to_float(row[idx] if idx < len(row) else None)
            if price is None:
                continue
            market = "日前" if "日前" in header else "实时" if "实时" in header else None
            if market:
                records.append({
                    "province": province,
                    "month": d[:7],
                    "datetime": f"{d}T{t}:00",
                    "market": market,
                    "price": price,
                    "volume": None,
                })
    return records, {
        "file": path.name,
        "sheet": ws.title,
        "priceColumns": {
            "日前": "电价类型=日前 → 节点均价" if type_col is not None else "日前节点均价或区域节点均价",
            "实时": "电价类型=实时 → 节点均价" if type_col is not None else "实时节点均价或区域节点均价",
        },
        "volumeColumns": {"日前": "典型光伏曲线", "实时": "典型光伏曲线"},
        "volumeSource": {
            "日前": "价格表无光伏字段，使用典型曲线独立权重",
            "实时": "价格表无光伏字段，使用典型曲线独立权重",
        },
        "records": len(records),
        "usesTypicalCurve": True,
    }


def ingest_horizontal_price_sheet(
    path: Path,
    province: str,
    typical_curve: dict[str, float],
) -> tuple[list[dict[str, Any]], dict[str, Any]] | None:
    wb = load_workbook_for_path(path)
    ws = wb[wb.sheetnames[0]]
    prepare_sheet(ws)
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [cell_text(v) for v in next(rows)]
    except StopIteration:
        return None

    date_col = find_col(headers, lambda h: h == "日期")
    type_col = find_col(headers, lambda h: h == "电价类型" or h == "类型")
    time_cols = [(idx, normalize_time(header)) for idx, header in enumerate(headers)]
    time_cols = [(idx, slot) for idx, slot in time_cols if slot is not None]
    if date_col is None or type_col is None or len(time_cols) < 12:
        return None

    if "24点" in path.name:
        return [], {
            "file": path.name,
            "sheet": ws.title,
            "ignored": "24点结算电价表不参与96点现货加权均价计算",
            "records": 0,
        }

    records: list[dict[str, Any]] = []
    for row in rows:
        d = normalize_date(row[date_col] if date_col < len(row) else None)
        market = market_from_text(row[type_col] if type_col < len(row) else None)
        if not d or not market:
            continue
        for idx, t in time_cols:
            price = to_float(row[idx] if idx < len(row) else None)
            if price is None:
                continue
            records.append({
                "province": province,
                "month": d[:7],
                "datetime": f"{d}T{t}:00",
                "market": market,
                "price": price,
                "volume": typical_curve.get(t, 0),
            })
    return records, {
        "file": path.name,
        "sheet": ws.title,
        "priceColumns": {"日前/实时": "横向96点电价"},
        "volumeColumns": {"日前": "典型光伏曲线", "实时": "典型光伏曲线"},
        "volumeSource": {
            "日前": "横向价格表无光伏字段，使用典型曲线独立权重",
            "实时": "横向价格表无光伏字段，使用典型曲线独立权重",
        },
        "records": len(records),
        "usesTypicalCurve": True,
    }


def extract_series(path: Path, province: str, typical_curve: dict[str, float]) -> tuple[
    list[dict[str, Any]],
    list[dict[str, Any]],
    dict[str, Any],
]:
    wb = load_workbook_for_path(path)
    ws = wb[wb.sheetnames[0]]
    prepare_sheet(ws)
    first_rows = []
    for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
        first_rows.append(row)
    headers, start_row = flatten_headers(first_rows)
    date_col = find_col(headers, lambda h: h == "日期" or h.endswith("-日期"))
    time_col = find_col(headers, is_time_header)
    if date_col is None or time_col is None:
        return [], [], {"file": path.name, "ignored": "未识别日期/时刻列"}
    type_col = find_col(headers, lambda h: h == "类型" or h.endswith("-类型") or h == "电价类型")
    inferred_file_market = infer_market_from_filename(path.name)
    volume_cols = select_volume_columns(headers, province)
    prices: list[dict[str, Any]] = []
    volumes: list[dict[str, Any]] = []
    used_typical_curve = any(v[2].endswith("典型光伏曲线") for v in volume_cols.values())
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        d = normalize_date(row[date_col] if date_col < len(row) else None)
        t = normalize_time(row[time_col] if time_col < len(row) else None)
        if not d or not t:
            continue
        row_market = market_from_text(row[type_col] if type_col is not None and type_col < len(row) else None)
        markets = [row_market] if row_market else [inferred_file_market] if inferred_file_market else ["日前", "实时"]
        for market in markets:
            value, header = row_price(row, headers, market, row_market or inferred_file_market, province)
            if value is not None:
                prices.append({
                    "province": province,
                    "month": d[:7],
                    "datetime": f"{d}T{t}:00",
                    "market": market,
                    "price": value,
                    "header": header,
                })
        for market in markets:
            if market not in volume_cols:
                continue
            idx, header, source = volume_cols[market]
            other_market = "实时" if market == "日前" else "日前"
            fallback_idx = volume_cols.get(other_market, (-1, "", ""))[0]
            if fallback_idx == idx:
                fallback_idx = -1
            value, used_typical = row_volume(
                row,
                idx,
                t,
                typical_curve,
                fallback_idx,
                province not in PROVINCES_WITHOUT_TYPICAL_VOLUME,
            )
            if used_typical:
                used_typical_curve = True
            if value is None:
                continue
            volumes.append({
                "province": province,
                "month": d[:7],
                "datetime": f"{d}T{t}:00",
                "market": market,
                "volume": value,
                "header": header,
                "source": source,
            })
    return prices, volumes, {
        "file": path.name,
        "sheet": ws.title,
        "priceColumns": selected_price_columns(headers, province, inferred_file_market),
        "volumeColumns": {k: v[1] for k, v in volume_cols.items()},
        "volumeSource": {k: v[2] for k, v in volume_cols.items()},
        "pricePoints": len(prices),
        "volumePoints": len(volumes),
        "usesTypicalCurve": used_typical_curve,
    }


def month_selected(month: str, years: set[int] | None, months: set[str] | None) -> bool:
    if months is not None:
        return month in months
    if years is not None:
        return int(month[:4]) in years
    return True


def path_may_match(path: Path, years: set[int] | None, months: set[str] | None) -> bool:
    file_month = month_from_filename(path.name)
    if file_month:
        return month_selected(file_month, years, months)
    path_years = {int(value) for value in re.findall(r"20\d{2}", path.name)}
    if months is not None and path_years:
        return int(next(iter(months))[:4]) in path_years
    if years is not None and path_years:
        return bool(path_years & years)
    return True


def calculate_slice(years: set[int] | None = None, months: set[str] | None = None) -> dict[str, Any]:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    typical_curves = load_typical_curves()
    common_curve = typical_curves.get("通用", {})
    params = load_params()
    provinces = [p for p in ROOT.iterdir() if p.is_dir() and not p.name.startswith(".") and p.name not in {"public", "scripts"}]
    output: dict[str, Any] = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "provinces": [],
        "monthly": [],
        "mappings": [],
        "params": {},
    }
    curve_needed: set[str] = set()

    def selected_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
        return [rec for rec in records if month_selected(rec["month"], years, months)]

    for province_dir in sorted(provinces, key=lambda p: p.name):
        province = province_dir.name
        typical_curve = typical_curves.get(province, common_curve)
        by_key: dict[tuple[str, str], dict[str, float]] = defaultdict(lambda: {"priceVolume": 0.0, "volume": 0.0, "points": 0})
        mapping_rows = []
        staged_prices: dict[tuple[str, str, str], float] = {}
        staged_volumes: dict[tuple[str, str, str], tuple[float, str]] = {}

        source_paths = [
            path for path in sorted(province_dir.rglob("*.xlsx"))
            if not path.name.startswith("~$") and path_may_match(path, years, months)
        ]
        for path in source_paths:
            file_month = month_from_filename(path.name)
            try:
                split_result = ingest_split_market_workbook(path, province, typical_curve)
                if split_result is not None:
                    records, info = split_result
                    records = selected_records(records)
                    info["records"] = len(records)
                    for rec in records:
                        key = (rec["month"], rec["market"])
                        by_key[key]["priceVolume"] += rec["price"] * rec["volume"]
                        by_key[key]["volume"] += rec["volume"]
                        by_key[key]["points"] += 1
                    mapping_rows.append(info)
                else:
                    horizontal_result = ingest_horizontal_price_sheet(path, province, typical_curve)
                    if horizontal_result is not None:
                        records, info = horizontal_result
                        records = selected_records(records)
                        info["records"] = len(records)
                        for rec in records:
                            key = (rec["month"], rec["market"])
                            by_key[key]["priceVolume"] += rec["price"] * rec["volume"]
                            by_key[key]["volume"] += rec["volume"]
                            by_key[key]["points"] += 1
                        if info.get("usesTypicalCurve"):
                            curve_needed.add(province)
                        mapping_rows.append(info)
                    elif "节点价格数据" in path.name or "_南方_价格_" in path.name:
                        price_records, info = ingest_southern_price_sheet(path, province)
                        price_records = selected_records(price_records)
                        info["records"] = len(price_records)
                        for rec in price_records:
                            staged_prices[(rec["datetime"], rec["market"], rec["month"])] = rec["price"]
                        mapping_rows.append(info)
                    else:
                        separate_price_patterns = SEPARATE_PRICE_FILE_RULES.get(province, ())
                        if any(pattern in path.name for pattern in separate_price_patterns):
                            prices, volumes, info = extract_series(path, province, typical_curve)
                            prices = selected_records(prices)
                            volumes = selected_records(volumes)
                            info["pricePoints"] = len(prices)
                            info["volumePoints"] = len(volumes)
                            for rec in prices:
                                staged_prices[(rec["datetime"], rec["market"], rec["month"])] = rec["price"]
                            for rec in volumes:
                                if "典型光伏曲线" not in rec["source"]:
                                    staged_volumes[(rec["datetime"], rec["market"], rec["month"])] = (rec["volume"], rec["source"])
                            info["volumeColumns"] = {
                                "日前": "关联供需/新能源负荷文件的日前光伏量",
                                "实时": "关联供需/新能源负荷文件的实时光伏量；缺失时复用日前光伏量",
                            }
                            info["volumeSource"] = {
                                "日前": "按日期时刻关联量文件；缺失时复用实时量，仍缺失则使用典型曲线",
                                "实时": "按日期时刻关联量文件；缺失时复用日前量，仍缺失则使用典型曲线",
                            }
                            info["usesTypicalCurve"] = False
                            mapping_rows.append(info)
                            continue
                        records, info = ingest_standard_sheet(path, province, typical_curve)
                        records = selected_records(records)
                        info["records"] = len(records)
                        for rec in records:
                            key = (rec["month"], rec["market"])
                            by_key[key]["priceVolume"] += rec["price"] * rec["volume"]
                            by_key[key]["volume"] += rec["volume"]
                            by_key[key]["points"] += 1
                        if info.get("usesTypicalCurve"):
                            curve_needed.add(province)
                        if not records:
                            prices, volumes, split_info = extract_series(path, province, typical_curve)
                            prices = selected_records(prices)
                            volumes = selected_records(volumes)
                            split_info["pricePoints"] = len(prices)
                            split_info["volumePoints"] = len(volumes)
                            for rec in prices:
                                staged_prices[(rec["datetime"], rec["market"], rec["month"])] = rec["price"]
                            for rec in volumes:
                                if "典型光伏曲线" not in rec["source"]:
                                    staged_volumes[(rec["datetime"], rec["market"], rec["month"])] = (rec["volume"], rec["source"])
                            if split_info.get("usesTypicalCurve"):
                                curve_needed.add(province)
                            info = split_info
                        mapping_rows.append(info)
            except Exception as exc:
                mapping_rows.append({"file": path.name, "error": f"{type(exc).__name__}: {exc}"})

        # Second pass for southern split price/volume files.
        for path in source_paths:
            if "市场供需" not in path.name and "供需数据" not in path.name and "新能源负荷" not in path.name:
                continue
            try:
                file_month = month_from_filename(path.name)
                _prices, volumes, info = extract_series(path, province, typical_curve)
                volumes = selected_records(volumes)
                for rec in volumes:
                    if "典型光伏曲线" not in rec["source"]:
                        staged_volumes[(rec["datetime"], rec["market"], rec["month"])] = (rec["volume"], rec["source"])
                if info.get("usesTypicalCurve"):
                    curve_needed.add(province)
            except Exception:
                continue
        for (dt, market, month), price in staged_prices.items():
            volume_tuple = staged_volumes.get((dt, market, month))
            if not volume_tuple:
                other_market = "实时" if market == "日前" else "日前"
                volume_tuple = staged_volumes.get((dt, other_market, month))
            if not volume_tuple:
                if province in PROVINCES_WITHOUT_TYPICAL_VOLUME:
                    continue
                volume_tuple = (typical_curve.get(dt[11:16], 0), "典型光伏曲线")
                curve_needed.add(province)
            volume, _source = volume_tuple
            key = (month, market)
            by_key[key]["priceVolume"] += price * volume
            by_key[key]["volume"] += volume
            by_key[key]["points"] += 1

        province_months = []
        for (month, market), agg in sorted(by_key.items()):
            if not agg["volume"]:
                continue
            weighted = agg["priceVolume"] / agg["volume"]
            row = {
                "province": province,
                "month": month,
                "market": market,
                "weightedAvg": weighted,
                "volume": agg["volume"],
                "points": agg["points"],
            }
            output["monthly"].append(row)
            province_months.append(month)

        output["provinces"].append({"name": province, "months": sorted(set(province_months))})
        output["mappings"].append({"province": province, "files": mapping_rows})
        matched = match_params(province, params)
        if matched:
            output["params"][province] = matched

    ensure_curve_sheets(curve_needed)
    output["curveNeeded"] = sorted(curve_needed)
    return output


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    with path.open("r", encoding="utf-8") as fh:
        return json.load(fh)


def merge_mapping_lists(
    existing: list[dict[str, Any]],
    update: list[dict[str, Any]],
    replace_months: set[str] | None = None,
) -> list[dict[str, Any]]:
    by_file: dict[tuple[str, str], dict[str, Any]] = {}
    for mapping in existing:
        province = mapping.get("province", "")
        for item in mapping.get("files", []):
            file_month = month_from_filename(item.get("file", ""))
            if replace_months and file_month in replace_months:
                continue
            by_file[(province, item.get("file", ""))] = item
    for mapping in update:
        province = mapping.get("province", "")
        for item in mapping.get("files", []):
            by_file[(province, item.get("file", ""))] = item
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for (province, _filename), item in by_file.items():
        grouped[province].append(item)
    return [
        {"province": province, "files": sorted(files, key=lambda item: item.get("file", ""))}
        for province, files in sorted(grouped.items())
    ]


def write_year_cache(
    cache_id: str,
    years: list[int],
    output: dict[str, Any],
    replace_months: set[str] | None = None,
) -> None:
    LOCAL_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    PUBLIC_YEARLY_DIR.mkdir(parents=True, exist_ok=True)
    local_path = LOCAL_CACHE_DIR / f"{cache_id}.json"
    public_path = PUBLIC_YEARLY_DIR / f"{cache_id}.json"
    if replace_months:
        existing = load_json(local_path, {})
        if not existing:
            public_existing = load_json(public_path, {})
            existing = {"monthly": public_existing.get("monthly", []), "mappings": []}
        kept_rows = [row for row in existing.get("monthly", []) if row.get("month") not in replace_months]
        output["monthly"] = kept_rows + output.get("monthly", [])
        output["mappings"] = merge_mapping_lists(
            existing.get("mappings", []),
            output.get("mappings", []),
            replace_months,
        )
    output["cacheId"] = cache_id
    output["years"] = years
    output["monthly"] = sorted(
        output.get("monthly", []),
        key=lambda row: (row["province"], row["month"], row["market"]),
    )
    with local_path.open("w", encoding="utf-8") as fh:
        json.dump(output, fh, ensure_ascii=False, indent=2)
    public_cache = {
        "cacheId": cache_id,
        "years": years,
        "generatedAt": output["generatedAt"],
        "monthly": output["monthly"],
    }
    with public_path.open("w", encoding="utf-8") as fh:
        json.dump(public_cache, fh, ensure_ascii=False, indent=2)
    print(f"Wrote yearly cache: {public_path}")


def merge_cached_outputs() -> dict[str, Any]:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    PUBLIC_YEARLY_DIR.mkdir(parents=True, exist_ok=True)
    rows_by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
    for path in sorted(PUBLIC_YEARLY_DIR.glob("*.json")):
        cache = load_json(path, {})
        for row in cache.get("monthly", []):
            rows_by_key[(row["province"], row["month"], row["market"])] = row
    monthly = sorted(rows_by_key.values(), key=lambda row: (row["province"], row["month"], row["market"]))
    params_source = load_params()
    province_names = sorted({row["province"] for row in monthly})
    params = {}
    provinces = []
    for province in province_names:
        province_months = sorted({row["month"] for row in monthly if row["province"] == province})
        provinces.append({"name": province, "months": province_months})
        matched = match_params(province, params_source)
        if matched:
            params[province] = matched
    public_output = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "provinces": provinces,
        "monthly": monthly,
        "params": params,
    }
    with (DATA_DIR / "dashboard-data.json").open("w", encoding="utf-8") as fh:
        json.dump(public_output, fh, ensure_ascii=False, indent=2)
    with (DATA_DIR / "dashboard-data.js").open("w", encoding="utf-8") as fh:
        fh.write("window.DASHBOARD_DATA = ")
        json.dump(public_output, fh, ensure_ascii=False)
        fh.write(";\n")

    mappings: list[dict[str, Any]] = []
    for path in sorted(LOCAL_CACHE_DIR.glob("*.json")):
        mappings = merge_mapping_lists(mappings, load_json(path, {}).get("mappings", []))
    if mappings:
        try:
            write_mapping_workbook(mappings)
        except PermissionError:
            print(f"Warning: {MAPPING_BOOK.name} is open; skipped mapping workbook refresh.")
            print("Close the workbook and run: python scripts/build_data.py merge")
    print(f"Merged {len(list(PUBLIC_YEARLY_DIR.glob('*.json')))} caches")
    print(f"Wrote {DATA_DIR / 'dashboard-data.json'}")
    print(f"Wrote {DATA_DIR / 'dashboard-data.js'}")
    return public_output


def build_history() -> None:
    output = calculate_slice(years={2024, 2025})
    write_year_cache("2024-2025", [2024, 2025], output)
    merge_cached_outputs()


def build_year(year: int) -> None:
    output = calculate_slice(years={year})
    write_year_cache(str(year), [year], output)
    merge_cached_outputs()


def build_month(month: str) -> None:
    if not re.fullmatch(r"20\d{2}-(0[1-9]|1[0-2])", month):
        raise ValueError("月份必须使用 YYYY-MM 格式，例如 2026-07")
    year = int(month[:4])
    output = calculate_slice(months={month})
    write_year_cache(str(year), [year], output, replace_months={month})
    merge_cached_outputs()


def print_cache_status() -> None:
    PUBLIC_YEARLY_DIR.mkdir(parents=True, exist_ok=True)
    caches = sorted(PUBLIC_YEARLY_DIR.glob("*.json"))
    if not caches:
        print("No yearly caches. Run: python scripts/build_data.py all --years 2026")
        return
    for path in caches:
        cache = load_json(path, {})
        months = sorted({row["month"] for row in cache.get("monthly", [])})
        period = f"{months[0]}..{months[-1]}" if months else "empty"
        print(f"{path.stem}: {len(cache.get('monthly', []))} rows, {period}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="按历史、年份或月份增量构建光伏看板数据")
    subparsers = parser.add_subparsers(dest="command")
    subparsers.add_parser("history", help="固化 2024-2025 历史数据")
    year_parser = subparsers.add_parser("year", help="重建指定年份，例如 year 2026")
    year_parser.add_argument("year", type=int)
    month_parser = subparsers.add_parser("month", help="只更新一个月份，例如 month 2026-07")
    month_parser.add_argument("month")
    subparsers.add_parser("merge", help="仅合并已有年度缓存，不读取各省原始数据")
    all_parser = subparsers.add_parser("all", help="重建历史数据和指定年份")
    all_parser.add_argument("--years", nargs="+", type=int, default=[2026])
    subparsers.add_parser("status", help="查看年度缓存覆盖范围")
    args = parser.parse_args()
    if args.command is None:
        args.command = "all"
        args.years = [2026]
    return args


def main() -> None:
    args = parse_args()
    if args.command == "history":
        build_history()
    elif args.command == "year":
        build_year(args.year)
    elif args.command == "month":
        build_month(args.month)
    elif args.command == "merge":
        merge_cached_outputs()
    elif args.command == "status":
        print_cache_status()
    elif args.command == "all":
        history_output = calculate_slice(years={2024, 2025})
        write_year_cache("2024-2025", [2024, 2025], history_output)
        for year in args.years:
            year_output = calculate_slice(years={year})
            write_year_cache(str(year), [year], year_output)
        merge_cached_outputs()


def join_mapping(values: Any) -> str:
    if not isinstance(values, dict):
        return ""
    return "\n".join(f"{key}: {value}" for key, value in values.items())


def load_existing_mapping_notes() -> dict[tuple[str, str], tuple[str, str]]:
    if not MAPPING_BOOK.exists():
        return {}
    wb = openpyxl.load_workbook(MAPPING_BOOK, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    notes = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        province = cell_text(row[0] if len(row) > 0 else "")
        filename = cell_text(row[1] if len(row) > 1 else "")
        if not province or not filename:
            continue
        manual = cell_text(row[10] if len(row) > 10 else "")
        instruction = cell_text(row[11] if len(row) > 11 else "")
        notes[(province, filename)] = (manual, instruction)
    return notes


def write_mapping_workbook(mappings: list[dict[str, Any]]) -> None:
    existing_notes = load_existing_mapping_notes()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "价量映射"
    headers = [
        "省份",
        "文件",
        "工作表",
        "价格字段",
        "光伏量字段",
        "量来源",
        "记录数",
        "价格点/量点",
        "是否使用典型曲线",
        "状态/备注",
        "人工确认",
        "维护说明",
    ]
    ws.append(headers)
    for mapping in mappings:
        province = mapping["province"]
        for item in mapping.get("files", []):
            status = item.get("error") or item.get("ignored") or ""
            manual, instruction = existing_notes.get((province, item.get("file", "")), ("", ""))
            if not instruction:
                instruction = "当前映射已按脚本识别结果更新；如需调整，请在本列补充字段与复用规则。"
            ws.append([
                province,
                item.get("file", ""),
                item.get("sheet", ""),
                join_mapping(item.get("priceColumns")),
                join_mapping(item.get("volumeColumns")),
                join_mapping(item.get("volumeSource")),
                item.get("records", ""),
                f"{item.get('pricePoints', '')}/{item.get('volumePoints', '')}" if "pricePoints" in item or "volumePoints" in item else "",
                "是" if item.get("usesTypicalCurve") else "否",
                status,
                manual,
                instruction,
            ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 10,
        "B": 42,
        "C": 28,
        "D": 36,
        "E": 36,
        "F": 32,
        "G": 12,
        "H": 14,
        "I": 16,
        "J": 28,
        "K": 16,
        "L": 58,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="DDEBF7")
    wb.save(MAPPING_BOOK)


if __name__ == "__main__":
    main()
